from ht_models import dict_sess, dict_engine, text, dict_base, Users, \
    Habits, UserHabitAssociations,UserHabitDays

from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

sess_users = dict_sess['sess_users']

def create_list_of_recorded_habits(formDict):
    habit_id_list = []

    for key,value in formDict.items():
        if key[:8] == 'checkbox':
            print("Key selected: ", key[9:])
            habit_id_list.append(key[9:])
    
    return habit_id_list

def get_user_habit_day_exists(user_id, habit_id, date):
    user_habit_day = sess_users.query(UserHabitDays).filter_by(
        user_id=user_id, habit_id = habit_id, date = date
    ).all()
    print('- in userHabitDayExists -')
    print("user_habit_day: ", user_habit_day)
    if len(user_habit_day) > 0:
        return True
    else:
        return False

def remove_table_name_from_df(df, table_name):
    # fix table names
    cols = list(df.columns)
    for col in cols:
        if col[:len(table_name)] == table_name:
            df = df.rename(columns=({col: col[len(table_name)+1:]}))
    
    return df

def create_user_habits_wb(df):
    # Initialize a workbook
    wb = Workbook()
    # Get the active sheet (the first one)
    ws = wb.active
    ws.title = "Data"

    # Transfer the data from df to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Format header row
    header_fill = PatternFill("solid", fgColor="D3D3D3") # Light grey fill
    header_border = Border(bottom=Side(border_style="thin")) # Underline
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = header_border
        cell.font = Font(bold=False)

    # Add a new sheet with the timestamp
    ws2 = wb.create_sheet(title="created on")
    ws2.append([datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    return wb